from django.shortcuts import render
from django.template import loader
from django.http import HttpResponse

from django.core import serializers
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

from django.contrib.auth.models import User
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.decorators import *
from json import JSONEncoder
from django.contrib.admin.views.decorators import staff_member_required
import json
from django.shortcuts import redirect



import random


from .models import *

from openpyxl import load_workbook



from django import forms

from io import BytesIO

class quizForm(forms.Form):
   file = forms.FileField()



@csrf_exempt
def addquiz(request):

    context = {}

    if request.method == "POST":
        quiz = Quiz.objects.get(id = request.POST['quiz'])
        file =request.FILES['file']
        file = file.read()
        wb = load_workbook(filename=BytesIO(file))
        ws = wb['quiz']
        for row in ws.rows:
            question = Question()
            for cell in row:
                if cell.value != None:
                    if cell.col_idx == 1 :
                        question = Question(Quiz=quiz , Text=cell.value)
                        question.save()
                    else :
                        answer = Answer(Question=question , Text=cell.value)
                        answer.save()




        context['status'] = "success"

        return JsonResponse(context, encoder=JSONEncoder)
    else:
        return render(request, 'addquiz.html', context)






@csrf_exempt
def addkey(request):

    context = {}

    if request.method == "POST":
        quiz = Quiz.objects.get(id = request.POST['quiz'])
        file =request.FILES['file']
        file = file.read()
        wb = load_workbook(filename=BytesIO(file))
        ws = wb['quiz']
        for row in ws.rows:
            question = Question()
            answer = Answer()
            for cell in row:
                if cell.value!= None :
                    if  cell.row == 1 :
                        continue
                    elif cell.col_idx == 1:
                        question = Question.objects.get(Number = cell.value , Quiz = quiz)
                    elif cell.col_idx ==2 :
                        answer = Answer.objects.get(Number = cell.value , Question=question)
                    else :
                        assessment = Assessment(Answer=answer ,Value=cell.value , Parameter=Parameter.objects.get(Name=
                                                                                                 ws.cell(column=cell.col_idx ,row=1).value  , Quiz=quiz))
                        assessment.save()





        context['status'] = "success"

        return JsonResponse(context, encoder=JSONEncoder)
    else:
        return render(request, 'addkey.html', context)





@csrf_exempt
def apiregister(request) :
    context = {}
    msg = request.POST
    context['status'] = 'failed'

    if 'name' not in msg or 'phonenumber' not in msg or 'password' not in msg or 'name' not in msg :
        context['message'] = 'مشخصات را به درستی وارد کنید'
        return JsonResponse(context, encoder=JSONEncoder)

    name = msg['name']
    Password = msg['password']
    Phonenumber = str(msg['phonenumber'])
    email = msg['email']

    if len(Phonenumber)!=10 or Phonenumber[0]!='9' or not Phonenumber.isdigit() :
        print(Phonenumber)
        context['message'] = 'شماره تلفن را به درستی وارد کنید.به صورت عدد انگلیسی با فرمت 9XXXXXXXX'
        return JsonResponse(context, encoder=JSONEncoder)

    if Student.objects.filter(PhoneNumber=Phonenumber).exists() or User.objects.filter(email=email).exists():
        context['message'] = 'مشخصات وارد شده تکراری میباشد.کاربری با این مشخصات قبلا ثبت نام کرده است.'
        return JsonResponse(context, encoder=JSONEncoder)


    user = User.objects.create_user(username=Phonenumber, password=Password , email=email)
    user.save()
    info = Student(Username=user,  PhoneNumber=Phonenumber, Name=name)
    info.save()

    context['message'] = 'ثبت نام شما با موفقیت انجام شد.'
    context['status'] = 'success'

    userauth = authenticate(request, username=Phonenumber, password=Password)
    login(request, userauth)

    return JsonResponse(context, encoder=JSONEncoder)


@csrf_exempt
def logout_view(request):
    logout(request)
    return redirect('/')



@csrf_exempt
def apilogin(request):
    context = {}
    if 'phonenumber' in request.POST and 'password' in request.POST:
        username = request.POST['phonenumber']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None :
            login(request, user)
            context['status'] = 'success'
            context['message'] = 'ورود با موفقیت انجام شد'
        else:
            context['message'] ='نام کاربری یا پسورد وارد شده اشتباه میباشد'
        return JsonResponse(context, encoder=JSONEncoder)

    print(request.POST)
    context['message'] = 'لطفا اطلاعات را به درستی وارد کنید'
    return JsonResponse(context, encoder=JSONEncoder)



def intro(request):
    context = {}
    if request.user.is_authenticated():
        student = Student.objects.get(user=request.user)
        state = student.State
        if state != 'result' :
            return render(request, 'intro.html', context)



    context = {}
    return render(request, 'intro.html', context)